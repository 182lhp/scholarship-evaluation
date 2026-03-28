"""
附件5：南京工业大学本科学生综合、单项奖学金获奖名单（Excel）
"""

import pandas as pd
from pathlib import Path

from config import LEVELS, LEVEL_ORDER, AWARD_AMOUNTS


def generate_award_list(
    ranked: pd.DataFrame,
    class_map: pd.Series,
    output_dir: str | Path,
) -> None:
    """生成附件5：南京工业大学本科学生综合、单项奖学金获奖名单.xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    award_levels = set(LEVELS)
    winners = ranked[ranked['综合等级'].isin(award_levels)].copy()
    winners['班级']    = winners['学号'].map(class_map).fillna('')
    winners['奖学金额'] = winners['综合等级'].map(AWARD_AMOUNTS).fillna(0).astype(int)
    winners['学号']    = winners['学号'].astype(str).str.replace(r'\.0$', '', regex=True)
    winners['_key']   = winners['综合等级'].map(LEVEL_ORDER).fillna(99)
    winners = (winners.sort_values(['_key', '班级', '综合排名'])
               .reset_index(drop=True))

    # ── openpyxl 样式 ──────────────────────────────────────────
    thin   = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font = Font(name='宋体', bold=True, size=18)
    hdr_font   = Font(name='宋体', bold=True, size=11)
    body_font  = Font(name='宋体', size=11)
    num_font   = Font(name='Times New Roman', size=11)

    TITLE   = '南京工业大学本科学生综合、单项奖学金获奖名单'
    HEADERS = ['序号', '学院', '班级', '学号', '姓名', '获奖等级', '奖学金额']
    COL_W   = [8, 40, 16, 16, 8, 10, 10]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = '获奖名单'

    # 标题行
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A1:G1')
    tc            = ws['A1']
    tc.value      = TITLE
    tc.font       = title_font
    tc.alignment  = center
    tc.border     = border

    # 表头行
    ws.row_dimensions[2].height = 22
    for ci, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        cell           = ws.cell(row=2, column=ci, value=h)
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border

    # 数据行
    for ri, (_, row) in enumerate(winners.iterrows(), start=3):
        ws.row_dimensions[ri].height = 22
        vals = [ri - 2, row.get('学院', ''), row.get('班级', ''),
                row['学号'], row['姓名'], row['综合等级'], row['奖学金额']]
        for ci, v in enumerate(vals, start=1):
            cell           = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = center
            cell.border    = border
            cell.font      = num_font if ci in (4, 7) else body_font

    fp = Path(output_dir) / '附件5：南京工业大学本科学生综合、单项奖学金获奖名单.xlsx'
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    wb.save(fp)
    print(f"\n> 📄 附件5 已生成：`{fp.name}`，共 **{len(winners)}** 名获奖学生\n")
