"""
Excel 格式化 + 保存工具
提供对 openpyxl workbook 进行统一样式处理的函数，
以及排名表 / 无资格名单的保存逻辑。
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from config import (
    LEVEL_COLORS, DISQ_COLOR, COL_WIDTHS, COL_WIDTH_DEFAULT,
    LEVEL_ORDER, SEMESTER_LABEL,
)
from utils.md import h2, md_table, safe_major_name

# ── 共享样式对象 ─────────────────────────────────────────────────
CENTER     = Alignment(horizontal='center', vertical='center')
FONT_NUM   = Font(name='Times New Roman')
FONT_TEXT  = Font(name='宋体')
FONT_HDR   = Font(name='宋体', bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def cell_font(value):
    """根据单元格值类型返回对应字体。"""
    if isinstance(value, (int, float)) and value == value:
        return FONT_NUM
    return FONT_TEXT


def _format_cell(cell, fill=None, id_col=False):
    """对单个单元格应用通用格式。"""
    if fill:
        cell.fill = fill
    cell.border    = THIN_BORDER
    cell.alignment = CENTER
    cell.font      = cell_font(cell.value)
    if id_col:
        cell.number_format = '@'
        cell.value = str(cell.value) if cell.value is not None else ''
        cell.font  = FONT_NUM


def _setup_worksheet(ws):
    """设置列宽并格式化表头行，返回学号列索引（1-based）或 None。"""
    for cell in ws[1]:
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS.get(
            cell.value, COL_WIDTH_DEFAULT
        )
    for cell in ws[1]:
        cell.font      = FONT_HDR
        cell.border    = THIN_BORDER
        cell.alignment = CENTER
    header_vals = [c.value for c in ws[1]]
    return header_vals.index('学号') + 1 if '学号' in header_vals else None


# ── 行染色 ───────────────────────────────────────────────────────

def apply_row_colors(filepath, df, level_col):
    """按 level_col 列的等级值对每行染色。"""
    wb = load_workbook(filepath)
    ws = wb.active
    if ws is None:
        return
    id_col_idx = _setup_worksheet(ws)
    for row_idx, level in enumerate(df[level_col], start=2):
        color = LEVEL_COLORS.get(level)
        fill  = PatternFill(fill_type='solid', fgColor=color) if color else None
        for cell in ws[row_idx]:
            is_id = id_col_idx is not None and cell.column == id_col_idx
            _format_cell(cell, fill=fill, id_col=is_id)
    wb.save(filepath)


def apply_row_colors_multi(filepath, sheet_dfs: dict, level_col: str):
    """对多 sheet 的 Excel 文件按 level_col 列批量染色。"""
    wb = load_workbook(filepath)
    for sheet_name, df in sheet_dfs.items():
        ws = wb[sheet_name]
        id_col_idx = _setup_worksheet(ws)
        for row_idx, level in enumerate(df[level_col], start=2):
            color = LEVEL_COLORS.get(level)
            fill  = PatternFill(fill_type='solid', fgColor=color) if color else None
            for cell in ws[row_idx]:
                is_id = id_col_idx is not None and cell.column == id_col_idx
                _format_cell(cell, fill=fill, id_col=is_id)
    wb.save(filepath)


def apply_all_red(filepath):
    """整张表所有数据行染无资格红色。"""
    wb = load_workbook(filepath)
    ws = wb.active
    if ws is None:
        return
    id_col_idx = _setup_worksheet(ws)
    fill = PatternFill(fill_type='solid', fgColor=DISQ_COLOR)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            is_id = id_col_idx is not None and cell.column == id_col_idx
            _format_cell(cell, fill=fill, id_col=is_id)
    wb.save(filepath)


# ── 保存排名结果 ─────────────────────────────────────────────────

def save_ranked_results(
    ranked_students: pd.DataFrame,
    merged_df: pd.DataFrame,
    class_map: pd.Series,
    disqualified: pd.DataFrame,
    output_dir: str,
) -> None:
    """保存排名表、成绩数据、无资格名单为 Excel 并应用格式。"""
    score_dir = Path(output_dir) / '成绩数据'
    rank_dir  = Path(output_dir) / '排名'
    score_dir.mkdir(parents=True, exist_ok=True)
    rank_dir.mkdir(parents=True, exist_ok=True)

    h2("保存结果")
    merged_df.to_excel(score_dir / f'{SEMESTER_LABEL}_合并成绩数据.xlsx', index=False)

    # 准备排名输出表
    DROP       = ['特等资格', '总学分绩点', '总学分']
    ranked_out = ranked_students.drop(
        columns=[c for c in DROP if c in ranked_students.columns]
    ).copy()
    ranked_out['_key'] = ranked_out['综合等级'].map(LEVEL_ORDER).fillna(99)
    ranked_out = (ranked_out
                  .sort_values(['专业', '_key', '综合排名'])
                  .drop(columns=['_key'])
                  .reset_index(drop=True))
    ranked_out['裸绩等级'] = ranked_out['裸绩等级'].replace('未获得奖学金', '')
    ranked_out['综合等级'] = ranked_out['综合等级'].replace('未获得奖学金', '')
    if '学号' in ranked_out.columns:
        ranked_out['学号'] = ranked_out['学号'].astype(str).str.replace(r'\.0$', '', regex=True)

    # 插入班级列（学号 → 班级，置于姓名之后）
    ranked_out['班级'] = ranked_out['学号'].map(class_map).fillna('')
    if '姓名' in ranked_out.columns:
        cols = ranked_out.columns.tolist()
        cols.remove('班级')
        insert_pos = cols.index('姓名') + 1
        cols.insert(insert_pos, '班级')
        ranked_out = ranked_out[cols]

    # 全校多 sheet
    fp        = rank_dir / f'{SEMESTER_LABEL}_校综合奖学金.xlsx'
    sheet_dfs = {}
    with pd.ExcelWriter(fp, engine='openpyxl') as writer:
        for major in sorted(ranked_out['专业'].dropna().unique()):
            md    = ranked_out[ranked_out['专业'] == major].copy()
            sname = safe_major_name(major)[:31]
            md.to_excel(writer, sheet_name=sname, index=False)
            sheet_dfs[sname] = md
    apply_row_colors_multi(fp, sheet_dfs, '综合等级')

    # 按专业单文件
    saved_rows = [[f'{SEMESTER_LABEL}_合并成绩数据.xlsx', len(merged_df), '成绩数据/']]
    saved_rows.append([fp.name, f"{len(sheet_dfs)} 个专业 sheet", '排名/'])
    for major in sorted(ranked_out['专业'].dropna().unique()):
        md   = ranked_out[ranked_out['专业'] == major].copy()
        path = rank_dir / f'{SEMESTER_LABEL}_校综合奖学金_{safe_major_name(major)}.xlsx'
        md.to_excel(path, index=False)
        apply_row_colors(path, md, '综合等级')
        saved_rows.append([path.name, len(md), '排名/'])

    # 无资格名单
    if len(disqualified) > 0:
        disq_out = disqualified.copy()
        if '学号' in disq_out.columns:
            disq_out['学号'] = disq_out['学号'].astype(str).str.replace(r'\.0$', '', regex=True)
        dp = Path(output_dir) / f'{SEMESTER_LABEL}_无资格名单.xlsx'
        disq_out.to_excel(dp, index=False)
        apply_all_red(dp)
        saved_rows.append([dp.name, f"{len(disq_out)} 条 / {disq_out['学号'].nunique()} 人", ''])

    md_table(['文件', '内容', '目录'], saved_rows, ['l', 'l', 'l'])
